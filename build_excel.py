#!/usr/bin/env python3
"""
Construye el Excel de informe a partir de los CSV de data/.

Uso:
    python build_excel.py                 # ultimos 7 dias
    python build_excel.py --dias 30
    python build_excel.py --desde 2026-08-01 --hasta 2026-08-15
"""

import argparse
import csv
from collections import OrderedDict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent / "data"
ARIAL = "Arial"
HORAS = list(range(8, 24))

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=ARIAL, size=14, bold=True, color="1F3864")
BODY = Font(name=ARIAL, size=10)
SMALL = Font(name=ARIAL, size=9, italic=True, color="7F7F7F")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def cargar(desde, hasta):
    filas = []
    for path in sorted(DATA_DIR.glob("*.csv")):
        try:
            dia = date.fromisoformat(path.stem)
        except ValueError:
            continue
        if not (desde <= dia <= hasta):
            continue
        with path.open(encoding="utf-8") as f:
            for r in csv.DictReader(f):
                if r.get("tipo_cola") not in ("STANDBY", "SINGLE_RIDER"):
                    continue
                filas.append(r)
    return filas


def cabecera(ws, fila, titulos, anchos):
    for i, t in enumerate(titulos, start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[fila].height = 30
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a


def hoja_datos(wb, filas):
    ws = wb.create_sheet("Datos")
    cabecera(ws, 1, ["Fecha", "Hora", "Minuto", "Parque", "Atraccion",
                     "Tipo de cola", "Espera (min)", "Estado"],
             [12, 7, 8, 24, 40, 15, 13, 14])
    r = 2
    for f in filas:
        espera = f.get("espera_min", "")
        ws.cell(row=r, column=1, value=f.get("fecha", ""))
        ws.cell(row=r, column=2, value=int(f["hora"]) if f.get("hora") else None)
        ws.cell(row=r, column=3, value=int(f["minuto"]) if f.get("minuto") else None)
        ws.cell(row=r, column=4, value=f.get("parque", ""))
        ws.cell(row=r, column=5, value=f.get("atraccion", ""))
        ws.cell(row=r, column=6, value=f.get("tipo_cola", ""))
        ws.cell(row=r, column=7, value=int(espera) if espera not in ("", None) else None)
        ws.cell(row=r, column=8, value=f.get("estado", ""))
        for c in range(1, 9):
            ws.cell(row=r, column=c).font = BODY
        r += 1
    ultima = r - 1
    ws.freeze_panes = "A2"
    if ultima >= 2:
        ws.auto_filter.ref = f"A1:H{ultima}"
    return ws, max(ultima, 2)


def hoja_resumen(wb, atracciones, n):
    ws = wb.create_sheet("Resumen", 0)
    ws["A1"] = "Tiempos de espera Disneyland Paris - resumen por atraccion"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Todas las cifras se calculan con formulas sobre la hoja Datos: al filtrar o "
                "ampliar el rango, se recalculan solas.")
    ws["A2"].font = SMALL

    cabecera(ws, 4, ["Atraccion", "Parque", "Media standby (min)", "Maximo standby (min)",
                     "Minimo standby (min)", "Media single rider (min)", "Muestras"],
             [40, 24, 16, 16, 16, 18, 11])

    D = f"Datos!$E$2:$E${n}"      # atraccion
    P = f"Datos!$D$2:$D${n}"      # parque
    T = f"Datos!$F$2:$F${n}"      # tipo de cola
    W = f"Datos!$G$2:$G${n}"      # espera

    r = 5
    for nombre, parque in atracciones:
        ws.cell(row=r, column=1, value=nombre)
        ws.cell(row=r, column=2, value=parque)
        cond = f'{D},$A{r},{P},$B{r},{T},"STANDBY"'
        ws.cell(row=r, column=3, value=f'=IFERROR(ROUND(AVERAGEIFS({W},{cond}),1),"")')
        ws.cell(row=r, column=4, value=f'=IFERROR(_xlfn.MAXIFS({W},{cond}),"")')
        ws.cell(row=r, column=5, value=f'=IFERROR(_xlfn.MINIFS({W},{cond}),"")')
        cond_sr = f'{D},$A{r},{P},$B{r},{T},"SINGLE_RIDER"'
        ws.cell(row=r, column=6, value=f'=IFERROR(ROUND(AVERAGEIFS({W},{cond_sr}),1),"")')
        ws.cell(row=r, column=7, value=f'=COUNTIFS({cond})')
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font, cell.border = BODY, BORDER
            if c > 2:
                cell.alignment = Alignment(horizontal="center")
            if r % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="EDF2F9")
        r += 1

    if r > 5:
        ws.auto_filter.ref = f"A4:G{r - 1}"
    ws.freeze_panes = "A5"
    return ws


def hoja_perfil(wb, atracciones, n):
    ws = wb.create_sheet("Perfil horario")
    ws["A1"] = "Espera media standby por franja de media hora"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Celda vacia = no hay muestras en esa franja en el rango cargado."
    ws["A2"].font = SMALL

    franjas = [(h, m) for h in HORAS for m in (0, 30)]
    cabecera(ws, 4, ["Atraccion"] + [f"{h:02d}:{m:02d}" for h, m in franjas],
             [40] + [7] * len(franjas))

    D = f"Datos!$E$2:$E${n}"
    H = f"Datos!$B$2:$B${n}"
    M = f"Datos!$C$2:$C${n}"
    T = f"Datos!$F$2:$F${n}"
    W = f"Datos!$G$2:$G${n}"

    r = 5
    for nombre, _parque in atracciones:
        ws.cell(row=r, column=1, value=nombre).font = BODY
        ws.cell(row=r, column=1).border = BORDER
        for i, (h, m) in enumerate(franjas, start=2):
            min_cond = f'{M},"<30"' if m == 0 else f'{M},">=30"'
            f = (f'=IFERROR(ROUND(AVERAGEIFS({W},{D},$A{r},{H},{h},{min_cond},'
                 f'{T},"STANDBY"),0),"")')
            c = ws.cell(row=r, column=i, value=f)
            c.font, c.border = BODY, BORDER
            c.alignment = Alignment(horizontal="center")
        r += 1
    ws.freeze_panes = "B5"
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dias", type=int, default=7)
    ap.add_argument("--desde")
    ap.add_argument("--hasta")
    ap.add_argument("--salida", default="tiempos_espera_dlp.xlsx")
    args = ap.parse_args()

    hasta = date.fromisoformat(args.hasta) if args.hasta else date.today()
    desde = date.fromisoformat(args.desde) if args.desde else hasta - timedelta(days=args.dias - 1)

    filas = cargar(desde, hasta)
    if not filas:
        print(f"No hay datos entre {desde} y {hasta}. Ejecuta collect.py primero.")
        return 1

    vistas = OrderedDict()
    for f in filas:
        vistas.setdefault((f["atraccion"], f["parque"]), None)
    atracciones = sorted(vistas.keys(), key=lambda x: (x[1], x[0]))

    wb = Workbook()
    wb.remove(wb.active)
    _, n = hoja_datos(wb, filas)
    hoja_resumen(wb, atracciones, n)
    hoja_perfil(wb, atracciones, n)
    wb.save(args.salida)
    print(f"{args.salida}: {len(filas)} muestras, {len(atracciones)} atracciones, "
          f"{desde} a {hasta}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
